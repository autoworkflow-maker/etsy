import os
from openpyxl import Workbook
from openpyxl.styles import Font
import random
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class ProductFileBuilder:
    THEMES = [
    {"name": "Midnight Blue", "primary": "1F4E78", "accent": "D9EAF7", "dark": "0F243E"},
    {"name": "Emerald Finance", "primary": "217346", "accent": "E2F0D9", "dark": "0B3D2E"},
    {"name": "Royal Purple", "primary": "5B2C83", "accent": "EADCF8", "dark": "2E1745"},
    {"name": "Charcoal Gold", "primary": "2F3437", "accent": "FFF2CC", "dark": "1C1C1C"},
    {"name": "Ocean Teal", "primary": "008C95", "accent": "DDEBF7", "dark": "004B50"},
    ]
    
    def __init__(self, folder):
        self.folder = folder
        os.makedirs(folder, exist_ok=True)
        

    def build_for_product(self, keyword, product_type):
        theme = random.choice(self.THEMES)
        text = f"{keyword} {product_type}".lower()

        if (
            "budget" in text
            or "tracker" in text
            or "planner" in text
            or "template" in text
        ):
            return self.build_budget_tracker(theme)

        return None

    def build_budget_tracker(self, theme):
        import os
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        filepath = os.path.join(self.folder, "budget_tracker.xlsx")

        wb = Workbook()

    # ── Transactions Sheet ─────────────────────────
    ws = wb.active
    ws.title = "Transactions"

    headers = [
        "Date",
        "Description",
        "Category",
        "Income",
        "Expense",
        "Balance"
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor=theme["primary"])
    white_font = Font(color="FFFFFF", bold=True)

    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    for row in range(2, 102):
        ws[f"A{row}"] = ""
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = ""

        ws[f"D{row}"] = 0
        ws[f"E{row}"] = 0

        if row == 2:
            ws[f"F{row}"] = "=D2-E2"
        else:
            ws[f"F{row}"] = f"=F{row-1}+D{row}-E{row}"

    widths = {
        "A": 15,
        "B": 30,
        "C": 20,
        "D": 15,
        "E": 15,
        "F": 18
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ── Dashboard Sheet ────────────────────────────
    dashboard = wb.create_sheet("Dashboard")

    dashboard["A1"] = "Budget Tracker Dashboard"
    dashboard["A1"].font = Font(
        bold=True,
        size=18,
        color=theme["dark"]
    )

    dashboard["A3"] = "Total Income"
    dashboard["B3"] = "=SUM(Transactions!D:D)"

    dashboard["A4"] = "Total Expenses"
    dashboard["B4"] = "=SUM(Transactions!E:E)"

    dashboard["A5"] = "Current Balance"
    dashboard["B5"] = "=B3-B4"

    dashboard["A7"] = "Instructions"
    dashboard["A8"] = "Enter transactions in the Transactions sheet."

    dashboard["A9"] = "Dashboard updates automatically."

    dashboard["A1"].fill = PatternFill(
        "solid",
        fgColor=theme["accent"]
    )

    dashboard.column_dimensions["A"].width = 28
    dashboard.column_dimensions["B"].width = 20

    # ── Categories Sheet ───────────────────────────
    categories = wb.create_sheet("Categories")

    categories["A1"] = "Suggested Categories"

    category_list = [
        "Income",
        "Salary",
        "Rent",
        "Groceries",
        "Utilities",
        "Insurance",
        "Transportation",
        "Fuel",
        "Entertainment",
        "Savings",
        "Investments",
        "Debt Payments",
        "Other"
    ]

    for i, cat in enumerate(category_list, start=2):
        categories[f"A{i}"] = cat

    categories.column_dimensions["A"].width = 30

    wb.save(filepath)

    return filepath
